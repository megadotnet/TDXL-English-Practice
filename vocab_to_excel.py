import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def parse_vocabulary(filepath):
    """解析词汇文本，提取单词、音标、词性、释义"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    words = []
    # 匹配格式: **word** /phonetic/ 或 **word** /phonetic, phonetic/
    # 然后匹配后面的词性和释义行
    pattern = re.compile(
        r'\*\*(.+?)\*\*\s+(/.+?/)\s*\n'
        r'(?:-\s+)?(.+)',
        re.MULTILINE
    )

    for match in pattern.finditer(content):
        word = match.group(1).strip()
        phonetic = match.group(2).strip()
        definition = match.group(3).strip()

        # 去掉音标的首尾斜杠
        phonetic = phonetic.strip('/')

        # 提取词性（第一个出现的词性标记）
        pos_match = re.match(r'^((?:n|v|vt|vi|a|ad|prep|conj|int|art|aux\.v|abbr|pron)\.?)', definition)
        pos = pos_match.group(1) if pos_match else ""

        # 释义（去掉前缀的词性标记和"- "）
        meaning = re.sub(r'^-\s*', '', definition).strip()

        words.append({
            'word': word,
            'phonetic': phonetic,
            'pos': pos,
            'meaning': meaning
        })

    return words


def create_excel(words, output_path):
    """将解析的词汇数据写入Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "英语词汇汇总表"

    # ---------- 样式定义 ----------
    header_font = Font(name='Microsoft YaHei', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    word_font = Font(name='Calibri', bold=True, size=11)
    phonetic_font = Font(name='Doulos SIL', size=10)
    pos_font = Font(name='Microsoft YaHei', size=10, color='1F4E79')
    meaning_font = Font(name='Microsoft YaHei', size=10)

    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7')
    )

    even_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    # ---------- 表头 ----------
    headers = ['序号', '单词', '音标', '词性', '中文释义']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ---------- 冻结首行 ----------
    ws.freeze_panes = 'A2'

    # ---------- 自动筛选 ----------
    ws.auto_filter.ref = f"A1:E1"

    # ---------- 数据写入 ----------
    for idx, w in enumerate(words, 1):
        row = idx + 1
        row_data = [idx, w['word'], w['phonetic'], w['pos'], w['meaning']]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            if col_idx == 1:  # 序号
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 2:  # 单词
                cell.font = word_font
            elif col_idx == 3:  # 音标
                cell.font = phonetic_font
            elif col_idx == 4:  # 词性
                cell.font = pos_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 5:  # 释义
                cell.font = meaning_font

            # 偶数行背景色（斑马纹）
            if idx % 2 == 0:
                cell.fill = even_fill

    # ---------- 列宽设置 ----------
    ws.column_dimensions['A'].width = 8    # 序号
    ws.column_dimensions['B'].width = 22   # 单词
    ws.column_dimensions['C'].width = 42   # 音标
    ws.column_dimensions['D'].width = 18   # 词性
    ws.column_dimensions['E'].width = 55   # 释义

    # ---------- 行高 ----------
    ws.row_dimensions[1].height = 25

    # ---------- 添加统计信息 Sheet ----------
    ws2 = wb.create_sheet("统计信息")
    stats = [
        ['统计项', '数值'],
        ['总词汇数', len(words)],
        ['名词(n.)', sum(1 for w in words if w['pos'] == 'n.')],
        ['动词(vt./vi./v.)', sum(1 for w in words if w['pos'] in ('vt.', 'vi.', 'v.'))],
        ['形容词(a.)', sum(1 for w in words if w['pos'] == 'a.')],
        ['副词(ad.)', sum(1 for w in words if w['pos'] == 'ad.')],
        ['介词(prep.)', sum(1 for w in words if w['pos'] == 'prep.')],
        ['连词(conj.)', sum(1 for w in words if w['pos'] == 'conj.')],
        ['代词(pron.)', sum(1 for w in words if w['pos'] == 'pron.')],
        ['其他', sum(1 for w in words if w['pos'] not in ('n.', 'vt.', 'vi.', 'v.', 'a.', 'ad.', 'prep.', 'conj.', 'pron.', ''))],
    ]
    for row in stats:
        ws2.append(row)

    wb.save(output_path)
    print(f"✅ 成功转换 {len(words)} 个词汇，已保存至: {output_path}")


# ========== 主程序 ==========
if __name__ == '__main__':
    # 将文档内容保存为 .txt 文件后运行
    input_file = 'TDXL英语单词汇总表.md'    # 输入文件路径
    output_file = '英语词汇汇总表.xlsx'  # 输出Excel路径

    words = parse_vocabulary(input_file)
    create_excel(words, output_file)
